#!/usr/bin/env python3
# -*- coding:utf-8 -*-
# @Author: ysy
# @Time: 2022-04-12 16:13

import re
import shutil
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from src.util.my_wx_auto import *
from src.util.bai_du_api_util import *
from src.util.common_util import *
from src.util.path_util import *
from src.util.excel_util import Xlsx

# 图片插入在Excel中的宽高
IMAGE_SIZE = (72, 156)
# 图片单元格的宽高
IMAGE_CELL_SIZE = (12, 120)

# 房间列表
ROOM_NUMBER = [104, 106, 108, 109, 111, 204, 206, 208, 209, 211, 213, 301, 306, 308, 309, 311, 313, 401, 406, 408, 409, 411, 413, 501, 506, 508, 509, 511, 513, 601, 604, 606, 608, 609, 611, 613, 701, 704, 706, 708, 709, 711, 713, 801, 806, 808, 809, 811, 813, 901, 904, 909, 911, 913, 1002, 1004, 1006, 1008, 1009, 1011, 1013, 1101, 1106, 1108, 1109, 1113, 1201, 1204, 1206, 1208, 1209, 1211, 1213, 1301, 1304, 1306, 1308, 1309, 1311, 1313, 1401, 1405, 1409, 1411, 1413, 1505, 1506, 1508, 1509, 1511, 1601, 1604, 1606, 1608, 1610, 1611, 1613, 1701, 1704, 1708, 1709, 1710]
# 昵称替换
NICKNAME_REPLACE_DICT = {'30弄2号': '', '30-2-': '', '8o1': '801', '张苹迦': '1004', '陈赛平': '1309', '黄丽云': '204', '匡长铭': '211', '沈永康': '508', '兰兰': '506', '石建兰': '608', '王学廉': '1008', '梧桐佬': '811', '笑咪咪': '1209', '许樑': '813', '勋': '613', '映山红': '808', '张菊香': '206', '张忠清': '209', '朱明华': '1002', '殷克忠': '308'}


def save_wx_picture(pic_save_dir, group_name, last_msg_size):
    """
        保存微信聊天图片
    :param pic_save_dir: 图片保存路径
    :param group_name: 微信的群名称
    :param last_msg_size: 从最后多少条消息
    :return: 返回保存的图片名list
    """
    wx = WeChat()
    wx.GetSessionList()
    wx.ChatWith(group_name, RollTimes=1)

    # 保存消息最后size中的图片
    msgItemArray = wx.GetMessageItemArray(size=last_msg_size)
    pic_name_list = list()
    for item in msgItemArray:
        click_result = wx.click_picture(item)
        if click_result:
            pic_name = WxUtils.SavePic(item, fileName='{date}-{nickName}-{random}', saveDirPath=pic_save_dir, nickname_replace_dict=NICKNAME_REPLACE_DICT)
            pic_name_list.append(pic_name)
    print('pic list size:', len(pic_name_list))
    return pic_name_list


def get_room_number(pic_name):
    """
        根据图片名获取房间号，将图片名使用'-'分隔取第二个，再替换掉非数字的就是房间号
    :param pic_name: 格式如{date}-{nickName}-{random}的图片名
    :return:
    """
    return re.sub('[^0-9]', '', pic_name.split('-')[1])


def get_valid_pic_path_list(pic_name_list, handle_base_dir, no_room_number_dir):
    """
        根据图片名获取房间号，并筛选出有效的
    :param pic_name_list: 图片名称列表
    :param handle_base_dir: 图片所在目录
    :param no_room_number_dir: 没有房间号的图片
    :return: （有效的图片路径列表(房间号, 图片全路径)，没有图片的房间号列表，不在房间号列表中的图片名(房间号, 图片名)）
    """
    valid_pic_path_list = []
    no_pic_room_number_list = list(ROOM_NUMBER)
    not_in_room_number_list = []
    for pic_name in pic_name_list:
        if is_image(pic_name):
            room_number = get_room_number(pic_name)
            if not room_number:
                print('pic_name:', pic_name, 'can\'t get room number, move to', no_room_number_dir)
                move_no_room_number_picture(pic_name, handle_base_dir, no_room_number_dir)
                continue
            room_number = int(room_number)
            if room_number not in ROOM_NUMBER:
                not_in_room_number_list.append((room_number, pic_name))
            if room_number in no_pic_room_number_list:
                no_pic_room_number_list.remove(room_number)
            valid_pic_path_list.append((int(room_number), os.path.join(handle_base_dir, pic_name)))
    valid_pic_path_list.sort(key=lambda x: x[0])
    return valid_pic_path_list, no_pic_room_number_list, not_in_room_number_list


def move_no_room_number_picture(pic_name, src_path, dst_path):
    if not os.path.exists(dst_path):
        os.mkdir(dst_path)
    move_to_path = os.path.join(dst_path, pic_name)
    shutil.move(os.path.join(src_path, pic_name), move_to_path)


def handle_result(result_xlsx, no_pic_room_number_list, not_in_room_number_list, skip_room_number_list):
    """
        打印结果，写入sheet
    :param result_xlsx:
    :param no_pic_room_number_list: 未发送图片的房间号
    :param not_in_room_number_list: 房间号不在列表中
    :param skip_room_number_list: 跳过的房间号
    """
    # 从未发送图片的房间号中移除跳过的房间号
    for skip_room_number in skip_room_number_list:
        if skip_room_number in no_pic_room_number_list:
            no_pic_room_number_list.remove(skip_room_number)
    result_xlsx.worksheet.cell(1, 1, '跳过的房间号：' + str(skip_room_number_list))
    print('skip_room_number_list:', skip_room_number_list)
    result_xlsx.worksheet.cell(2, 1, '未发送图片的房间号：' + str(no_pic_room_number_list))
    print('no_pic_room_number_list:', no_pic_room_number_list)
    result_xlsx.worksheet.cell(3, 1, '房间号不在列表中：' + str(not_in_room_number_list))
    print('not_in_room_number_list:', not_in_room_number_list)


class HandleAntigenPicture(object):
    '''将微信聊天图片，保存到excel中，并将群昵称处理为房间号'''

    # 表头
    HEADER_ROW = ['房间号', '抗原检测图片']
    
    def __init__(self, handle_base_dir, result_excel_name='antigen_result.xlsx') -> None:
        self.handle_base_dir = handle_base_dir
        self.no_room_number_dir = os.path.join(handle_base_dir, 'no_room_number')
        self.result_xlsx = Xlsx(os.path.join(handle_base_dir, result_excel_name))
        self.result_xlsx.init_header(self.HEADER_ROW, {'B': IMAGE_CELL_SIZE[0]}, header_row_index=4)

    def save_to_excel(self, pic_name_list=None, skip_room_number_list=[]):
        pic_name_list = pic_name_list if pic_name_list else os.listdir(self.handle_base_dir)
        valid_pic_path_list, no_pic_room_number_list, not_in_room_number_list = get_valid_pic_path_list(pic_name_list, self.handle_base_dir, self.no_room_number_dir)
        for i, (room_number, pic_path) in enumerate(valid_pic_path_list):
            self.result_xlsx.write_row([room_number], row_height=IMAGE_CELL_SIZE[1])
            self.result_xlsx.add_image(pic_path, 'B', image_size=IMAGE_SIZE)
        handle_result(self.result_xlsx, no_pic_room_number_list, not_in_room_number_list, skip_room_number_list)
        self.result_xlsx.save()

    def run(self, group_name, last_msg_size):
        pic_name_list = save_wx_picture(self.handle_base_dir, group_name, last_msg_size)
        # pic_name_list = ['20220413-409室-7003.jpg']
        print(pic_name_list)
        self.save_to_excel(pic_name_list)


class HandleNucleicPicture(object):
    '''使用ocr识别核酸检测结果的截图，并保存到excel中，并将群昵称处理为房间号'''

    # ocr结果中必须有的词语，若没有表示非核酸检测结果截图
    OCR_HAS_WORDS_LIST = ['姓名', '证件号码', '核酸']
    # 表头
    HEADER_ROW = ['房间号', '姓名', '是否阴性', '核酸结果文字', '核酸截图']

    def __init__(self, handle_base_dir, result_excel_name='nucleic_result.xlsx') -> None:
        self.handle_base_dir = handle_base_dir
        self.no_room_number_dir = os.path.join(handle_base_dir, 'no_room_number')
        self.bai_du_api = BaiDuApi('g1S8iIbqBGGe8xKlAmWEC6DA', 'W3Kwt7XrN72PWXdsFhfDLMjoIL99D9TS')
        self.result_xlsx = Xlsx(os.path.join(handle_base_dir, result_excel_name))
        self.result_xlsx.init_header(self.HEADER_ROW, {'D': 12, 'E': IMAGE_CELL_SIZE[0]})

    def picture_ocr_handle(self, pic_name):
        ocr_result_list = self.bai_du_api.ocr(os.path.join(self.handle_base_dir, pic_name))
        # 若ocr结果中含有以下文字表示是核酸检测的截图
        if all_in_list(HandleNucleicPicture.OCR_HAS_WORDS_LIST, ocr_result_list):
            # 获取'姓名'的后一个词语为姓名
            name = ocr_result_list[ocr_result_list.index('姓名') + 1]
            # 获取'检测结'的后一个词语，且'阴'在其中则为true。因有的使用了大字版的微信，“检测结果”一行显示不完就成了“检测结”
            result_word = ocr_result_list[any_in_list(['检测结'], ocr_result_list)[1] + 1]
            return name, '阴' in result_word, result_word
        else:
            return None, None, None

    def save_to_excel(self, pic_name_list=None, skip_room_number_list=[]):
        pic_name_list = pic_name_list if pic_name_list else os.listdir(self.handle_base_dir)
        valid_pic_path_list, no_pic_room_number_list, not_in_room_number_list = get_valid_pic_path_list(pic_name_list, self.handle_base_dir, self.no_room_number_dir)
        for i, (room_number, pic_path) in enumerate(valid_pic_path_list):
            name, result, result_word = self.picture_ocr_handle(pic_path)
            # name, result, result_word = ('周仲韵<', True, '【阴性】')
            if not name:
                print('pic_path:', pic_path, 'is not a covid-19 test picture')
                continue
            self.result_xlsx.write_row([room_number, name, result, result_word], row_height=IMAGE_CELL_SIZE[1])
            self.result_xlsx.add_image(pic_path, 'E', image_size=IMAGE_SIZE)
            if not result:
                print('pic_path:', pic_path, 'name: ', name, 'is not negative, result_word:', result_word)
        handle_result(self.result_xlsx, no_pic_room_number_list, not_in_room_number_list, skip_room_number_list)
        self.result_xlsx.save()

    def run(self, group_name, last_msg_size):
        pic_name_list = save_wx_picture(self.handle_base_dir, group_name, last_msg_size)
        # pic_name_list = ['20220413-409室-7003.jpg']
        print(pic_name_list)
        self.save_to_excel(pic_name_list)


if __name__ == '__main__':
    # print('30弄2号' in '30弄2号111室')
    handle_base_dir = r'C:\data\wx-get-picture\KangYuanJieGuo\20220421'
    save_wx_picture(handle_base_dir, '30弄2号楼业主群', 900)
    # handleAntigenPicture = HandleAntigenPicture(handle_base_dir, result_excel_name='30弄2号楼-抗原统计-20220421.xlsx')
    # handleAntigenPicture.save_to_excel(skip_room_number_list=[1109, 1211, 1213, 1704])
    # handleAntigenPicture.run('30弄2号楼业主群', 600)

    # handleNucleicPicture = HandleNucleicPicture(r'C:\data\wx-get-picture\HeSuanJieGuo-test')
    # handleNucleicPicture.save_to_excel()
    # handleNucleicPicture.run('管一二街坊清美', 50)

    pass
