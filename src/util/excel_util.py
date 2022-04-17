#!/usr/bin/env python3
# -*- coding:utf-8 -*-
# @Author: ysy
# @Time: 2022-04-17 10:57


import os
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from src.util.logger_util import logger


class Xlsx(object):

    def __init__(self, file_path, is_create=True, sheet_index=0) -> None:
        self.file_path = file_path
        if os.path.exists(file_path) and not is_create:
            logger.info('load exists xlsx from: ' + self.file_path)
            self.workbook = load_workbook(file_path)
        else:
            logger.info('create xlsx will save to: ' + self.file_path)
            self.workbook = Workbook()
        self.worksheet = self.workbook.worksheets[sheet_index]
        self.next_write_row_index = 1

    def init_header(self, header_row: list, header_width_dict: dict):
        logger.info('init header row: {0}, width: {1}'.format(header_row, header_width_dict))
        for k, v in header_width_dict.items():
            self.worksheet.column_dimensions[k].width = v
        for i, header in enumerate(header_row):
            self.worksheet.cell(1, i + 1, header)
        self.next_write_row_index = 2

    def set_next_write_row_index(self, next_write_row_index):
        self.next_write_row_index = next_write_row_index

    def write_row(self, row: list, row_index=None, row_height=None):
        if not row_index:
            row_index = self.next_write_row_index
        logger.info('write row({0}): {1}'.format(row_index, row))
        if row_height:
            self.worksheet.row_dimensions[row_index].height = row_height
        for i, v in enumerate(row):
            self.worksheet.cell(row_index, i + 1, v)
        # 自动增长下一次写的行号
        self.next_write_row_index = row_index + 1

    def add_image(self, anchor, image_path, image_size=None):
        logger.info('add image({0}): {1}'.format(anchor, image_path))
        image = Image(image_path)
        image.width, image.height = image_size
        self.worksheet.add_image(image, anchor=anchor)

    def save(self):
        logger.info('save xlsx: ' + self.file_path)
        self.workbook.save(self.file_path)
        self.workbook.close()


if __name__ == '__main__':
    pass
